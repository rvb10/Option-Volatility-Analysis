from pathlib import Path
from datetime import datetime, date as Date, timedelta

from openpyxl import load_workbook

def _project_root() -> Path:
    return Path(__file__).resolve().parents[1]

def _parse_trade_date(trade_date_str: str) -> Date:
    return datetime.strptime(trade_date_str, "%Y-%m-%d").date()

def _to_date(x) -> Date | None:
    if isinstance(x, Date) and not isinstance(x, datetime):
        return x
    if isinstance(x, datetime):
        return x.date()
    if isinstance(x, str):
        x = x.strip()
        for fmt in ("%m/%d/%y", "%m/%d/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(x, fmt).date()
            except ValueError:
                continue
    return None

def _to_float(x) -> float | None:
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    if isinstance(x, str):
        try:
            # handle possible % signs
            x_clean = x.replace("%", "").strip()
            return float(x_clean)
        except ValueError:
            return None
    return None

def _get_index_info(trade_date: Date, raw_dir: Path):
    path = raw_dir / "Index Data.xlsx"
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    best_date = None
    best_S0 = None
    best_r = None
    best_q = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_date = row[0]  
        d = _to_date(raw_date)
        if d is None or d > trade_date:
            continue

        S0 = _to_float(row[1])       
        sofr = _to_float(row[5])     
        div_yield = _to_float(row[9]) 

        if S0 is None or sofr is None or div_yield is None:
            continue

        if best_date is None or d > best_date:
            best_date = d
            best_S0 = S0
            best_r = sofr / 100.0
            best_q = div_yield / 100.0

    if best_date is None:
        raise ValueError(
            f"No complete index row (SPY, SOFR, div yield) found on or before {trade_date}."
        )

    return {
        "S0": best_S0,
        "r": best_r,
        "q": best_q,
    }

def _get_iv_info(trade_date: Date, raw_dir: Path):
    path = raw_dir / "IV.xlsx"
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header = [str(c).strip() if c is not None else "" for c in header_row]
    col_idx = {name: i for i, name in enumerate(header)}

    required_cols = [
        "Date",
        "95% Moneyness Vol 1m",
        "105% Moneyness Vol 1m",
    ]
    for col in required_cols:
        if col not in col_idx:
            raise ValueError(f"Column '{col}' not found in IV.xlsx header: {header}")

    best_row = None
    best_date = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        d = _to_date(row[col_idx["Date"]])
        if d is None:
            continue
        if d <= trade_date and (best_date is None or d > best_date):
            best_date = d
            best_row = row

    if best_row is None:
        raise ValueError(f"No IV row found on or before {trade_date}")

    put_iv_pct = _to_float(best_row[col_idx["95% Moneyness Vol 1m"]])
    call_iv_pct = _to_float(best_row[col_idx["105% Moneyness Vol 1m"]])

    if put_iv_pct is None or call_iv_pct is None:
        raise ValueError("Failed to parse IVs from IV.xlsx.")

    return {
        "iv_put_entry": put_iv_pct / 100.0,   
        "iv_call_entry": call_iv_pct / 100.0,
    }


def _get_expiry_and_strikes(trade_date: Date, S0: float, raw_dir: Path):
    path = raw_dir / "options_parsed.xlsx"
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header = [str(c).strip() if c is not None else "" for c in header_row]
    col_idx = {name: i for i, name in enumerate(header)}

    needed = ["expiry", "option_type", "strike"]
    for col in needed:
        if col not in col_idx:
            raise ValueError(f"Column '{col}' not found in options_parsed.xlsx header: {header}")

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        expiry_raw = row[col_idx["expiry"]]
        expiry_date = _to_date(expiry_raw)
        if expiry_date is None:
            continue

        opt_type = row[col_idx["option_type"]]
        strike = _to_float(row[col_idx["strike"]])

        if strike is None or opt_type is None:
            continue

        records.append({
            "expiry": expiry_date,
            "option_type": str(opt_type).upper().strip(),
            "strike": strike,
        })

    if not records:
        raise ValueError("No valid option rows found in options_parsed.xlsx.")

    future_expiries = sorted({r["expiry"] for r in records if r["expiry"] > trade_date})
    if future_expiries:
        expiry_date = future_expiries[0]
    else:
        all_expiries = sorted({r["expiry"] for r in records})
        expiry_date = all_expiries[0]

    K_put_target = 0.95 * S0
    K_call_target = 1.05 * S0

    rec_at_exp = [r for r in records if r["expiry"] == expiry_date]

    puts = [r for r in rec_at_exp if r["option_type"] == "P"]
    calls = [r for r in rec_at_exp if r["option_type"] == "C"]

    if not puts:
        raise ValueError(f"No puts found for expiry {expiry_date} in options_parsed.xlsx.")

    K_put = min(puts, key=lambda r: abs(r["strike"] - K_put_target))["strike"]

    if calls:
        K_call = min(calls, key=lambda r: abs(r["strike"] - K_call_target))["strike"]
    else:
        K_call = round(K_call_target / 5.0) * 5.0

    return {
        "expiry_date": expiry_date,
        "K_put": K_put,
        "K_call": K_call,
    }

def get_inputs(trade_date_str: str = "2025-04-09") -> dict:
    root = _project_root()
    raw_dir = root / "data" / "raw"
    trade_date = _parse_trade_date(trade_date_str)

    idx_info = _get_index_info(trade_date, raw_dir)
    S0 = idx_info["S0"]
    r = idx_info["r"]
    q = idx_info["q"]

    iv_info = _get_iv_info(trade_date, raw_dir)
    iv_put_entry = iv_info["iv_put_entry"]
    iv_call_entry = iv_info["iv_call_entry"]

    opt_info = _get_expiry_and_strikes(trade_date, S0, raw_dir)
    expiry_date = opt_info["expiry_date"]
    K_put = opt_info["K_put"]
    K_call = opt_info["K_call"]

    valuation_date = expiry_date - timedelta(days=7)
    if valuation_date < trade_date:
        valuation_date = trade_date

    contract_multiplier = 100
    contracts_put_sold = 1000
    contracts_call_bought = 1000

    return {
        "S0": S0,
        "r": r,
        "q": q,
        "trade_date": trade_date,
        "expiry_date": expiry_date,
        "valuation_date": valuation_date,
        "K_put": K_put,
        "K_call": K_call,
        "iv_put_entry": iv_put_entry,
        "iv_call_entry": iv_call_entry,
        "contract_multiplier": contract_multiplier,
        "contracts_put_sold": contracts_put_sold,
        "contracts_call_bought": contracts_call_bought,
    }


if __name__ == "__main__":
    info = get_inputs("2025-04-09")
    print("=== Extracted Inputs ===")
    for k, v in info.items():
        print(f"{k}: {v}")
